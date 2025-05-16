import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import sys

def count_tags(xml_file):
    tag_counts = {}

    for event, element in ET.iterparse(xml_file, events=('start',)):
        tag_type = element.tag.split('/')[-1]
        tag_counts[tag_type] = tag_counts.get(tag_type, 0) + 1
        element.clear()

    return tag_counts

def write_report(tag_counts, output_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Tag", "Getelde aantal"])
    for tag_type, count in tag_counts.items():
        ws.append([tag_type, count])
    wb.save(output_file)

def compare_excel_files(file1, file2, output_file):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    ws1 = wb1.active
    ws2 = wb2.active

    compared_data = {}

    for row1, row2 in zip(ws1.iter_rows(min_row=2, values_only=True), ws2.iter_rows(min_row=2, values_only=True)):
        tag1, count1 = row1[0], int(row1[1])
        tag2, count2 = row2[0], int(row2[1])
        difference = count2 - count1
        compared_data[tag1] = (count1, count2, count1 == count2, difference)

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["Tag", f"Getelde aantal ({os.path.basename(file1)})", f"Getelde aantal ({os.path.basename(file2)})",
                "Gelijk?", "Verschil"])

    for tag, data in compared_data.items():
        ws3.append([tag, *data])

    wb3.save(output_file)

if __name__ == "__main__":
    # XML-bestanden en uitvoerbestanden
    xml_file1 = "\\owwnlms601\LDV_Export\check\I_ldv_cho224.xml"
    xml_file2 = "\\owwnlms601\LDV_Export\check\I_ldv_cho234.xml"
    output_excel1 = "\\owwnlms601\LDV_Export\check\excel1.xlsx"
    output_excel2 = "\\owwnlms601\LDV_Export\check\excel2.xlsx"
    today_date = datetime.now().strftime("%Y%m%d")
    comparison_output = f"\\owwnlms601\LDV_Export\check\rapport_vergelijk_LDV_xml_{today_date}.xlsx"

    # Uitvoeren van de functies met opgegeven bestanden
    tag_counts1 = count_tags(xml_file1)
    write_report(tag_counts1, output_excel1)

    tag_counts2 = count_tags(xml_file2)
    write_report(tag_counts2, output_excel2)

    compare_excel_files(output_excel1, output_excel2, comparison_output)
