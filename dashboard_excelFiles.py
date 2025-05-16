import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule
from openpyxl.drawing.image import Image

# Directory waar de Excel-bestanden staan
directory = r'C:\Users\linkeddata\Downloads\I_ldv_cho.xml\analyse'

# Lege lijst om dataframes op te slaan
dfs = []

# Functie om de datum uit de bestandsnaam te extraheren en om te zetten in een leesbaar formaat
def extract_date_from_filename(filename):
    try:
        # Assuming the date format is always at the end before the extension, e.g., "I_ldv_cho_0608.xlsx"
        date_str = filename.split('_')[-1].split('.')[0]
        date_obj = datetime.strptime(f'2024{date_str}', '%Y%d%m')  # Aanname: het jaar is 2024
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        print(f"Bestandsnaam {filename} bevat geen geldige datum in het formaat DDMM.")
        return None

# Maak een nieuw Excel-werkboek aan
wb = Workbook()

# Voeg uitleg toe aan een nieuw tabblad
explanation_ws = wb.active
explanation_ws.title = "Uitleg"
explanation_text = [
    "Uitleg van de Analyse:",
    "",
    "1. Verschil in Aantal per Tag:",
    "   - Deze tabel toont het verschil in aantal per tag tussen verschillende datums.",
    "   - Elke waarde geeft het verschil weer tussen de huidige en de vorige periode.",
    "   - Bijvoorbeeld, een waarde van -40 betekent dat er 40 minder items zijn vergeleken met de vorige periode.",
    "",
    "2. Gemiddelde en Standaarddeviatie:",
    "   - Het tabblad 'Gemiddelde_StDev' toont het gemiddelde en de standaarddeviatie van de aantallen voor elke tag over de gehele periode.",
    "   - Het gemiddelde geeft de centrale tendens van de data weer.",
    "   - De standaarddeviatie toont de spreiding of variabiliteit van de data.",
    "",
    "3. Grafieken:",
    "   - De grafieken op het 'Dashboard' en 'Gemiddelde_StDev' tabblad geven een visuele representatie van de trends en statistieken.",
    "   - Gebruik de grafieken om snel inzicht te krijgen in de veranderingen en variabiliteit in de data.",
    "",
    "Als er nog vragen zijn over deze analyse, neem dan contact op met de gegevensbeheerder."
]

# Voeg de uitlegtekst toe aan rijen
for i, line in enumerate(explanation_text, 1):
    explanation_ws[f"A{i}"] = line

# Voeg het Dashboard tabblad toe
summary_ws = wb.create_sheet(title="Dashboard")

# Doorloop alle bestanden in de directory
for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)
        # Extract the date from the filename
        period = extract_date_from_filename(filename)
        if period is None:
            continue  # Sla het bestand over als er geen geldige datum is
        
        # Lees de eerste sheet in (algemene benadering voor de sheet naam)
        df = pd.read_excel(file_path, sheet_name=0)
        df.columns = df.columns.map(str)  # Zorg dat alle kolomkoppen strings zijn
        df['Period'] = period  # Voeg een kolom toe voor de periode
        dfs.append(df)
        
        # Voeg de data toe aan een nieuw werkblad
        ws = wb.create_sheet(title=period)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

# Combineer alle dataframes
combined_df = pd.concat(dfs, ignore_index=True)

# Bereken de verschillen tussen de periodes
combined_df['Previous_Count'] = combined_df.groupby('Tag')['Count'].shift(1)
combined_df['Difference'] = combined_df['Count'] - combined_df['Previous_Count']

# Filter de eerste periode uit omdat er geen verschil kan worden berekend
combined_df = combined_df.dropna(subset=['Difference'])

# Groepeer de data op 'Tag' en 'Period' en som de 'Difference' kolom
trend_data = combined_df.pivot_table(values='Difference', index='Tag', columns='Period', aggfunc='sum', fill_value=0)

# Zorg dat alle kolomkoppen strings zijn
trend_data.columns = trend_data.columns.map(str)

# Voeg de gecombineerde data toe aan het overzichtstabblad
for r in dataframe_to_rows(trend_data, index=True, header=True):
    summary_ws.append(r)

# Maak een grafiek van de trends
chart = LineChart()
chart.title = "Trends over Tijd per Tag"
chart.style = 13
chart.y_axis.title = 'Verschil in Aantal'
chart.x_axis.title = 'Periode'

# Data voor de grafiek
data = Reference(summary_ws, min_col=2, min_row=1, max_row=len(trend_data.index)+1, max_col=len(trend_data.columns)+1)
categories = Reference(summary_ws, min_col=1, min_row=2, max_row=len(trend_data.index)+1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
summary_ws.add_chart(chart, "E5")

# Extra analyses
# Percentage Wijziging
combined_df['Percentage Change'] = (combined_df['Difference'] / combined_df['Previous_Count']) * 100

# Gemiddelde en Standaarddeviatie
avg_std_data = combined_df.groupby('Tag')['Count'].agg(['mean', 'std']).reset_index()
avg_std_data.columns = avg_std_data.columns.map(str)  # Zorg dat alle kolomkoppen strings zijn

# Voeg de gemiddelde en standaarddeviatie gegevens toe aan een nieuw tabblad
avg_std_ws = wb.create_sheet(title="Gemiddelde_StDev", index=2)
for r in dataframe_to_rows(avg_std_data, index=False, header=True):
    avg_std_ws.append(r)

# Maak grafieken voor gemiddelde en standaarddeviatie
avg_chart = BarChart()
avg_chart.title = "Gemiddelde per Tag"
avg_chart.style = 13
avg_chart.y_axis.title = 'Gemiddelde'
avg_chart.x_axis.title = 'Tag'

data = Reference(avg_std_ws, min_col=2, min_row=1, max_row=len(avg_std_data.index)+1, max_col=2)
categories = Reference(avg_std_ws, min_col=1, min_row=2, max_row=len(avg_std_data.index)+1)

avg_chart.add_data(data, titles_from_data=True)
avg_chart.set_categories(categories)
avg_std_ws.add_chart(avg_chart, "E5")

std_chart = BarChart()
std_chart.title = "Standaarddeviatie per Tag"
std_chart.style = 13
std_chart.y_axis.title = 'Standaarddeviatie'
std_chart.x_axis.title = 'Tag'

data = Reference(avg_std_ws, min_col=3, min_row=1, max_row=len(avg_std_data.index)+1, max_col=3)
categories = Reference(avg_std_ws, min_col=1, min_row=2, max_row=len(avg_std_data.index)+1)

std_chart.add_data(data, titles_from_data=True)
std_chart.set_categories(categories)
avg_std_ws.add_chart(std_chart, "E20")

# Voeg een draaitabel toe
pivot_ws = wb.create_sheet(title="Draaitabel")
pivot_df = combined_df.pivot_table(values='Count', index='Tag', columns='Period', aggfunc='sum', fill_value=0)
for r in dataframe_to_rows(pivot_df, index=True, header=True):
    pivot_ws.append(r)

# Voeg filters toe aan de draaitabel
pivot_ws.auto_filter.ref = pivot_ws.dimensions

# Voeg een gauge toe (meterdiagram)
gauge_ws = wb.create_sheet(title="Gauge")
gauge_ws["A1"] = "Gemiddelde Count per Tag"
gauge_avg_df = combined_df.groupby('Tag')['Count'].mean().reset_index()
for r in dataframe_to_rows(gauge_avg_df, index=False, header=True):
    gauge_ws.append(r)

# Voeg databars toe als visualisatie van de gauge
databar_rule = DataBarRule(start_type="num", start_value="0", end_type="num", end_value="100",
                           color="FF638EC6", showValue="None")
gauge_ws.conditional_formatting.add(f"B2:B{len(gauge_avg_df)+1}", databar_rule)


# Sla het werkboek op in de analyse map met de datum toegevoegd aan de bestandsnaam
current_date = datetime.now().strftime('%Y%m%d')
output_path = os.path.join(directory, f'combined_analysis_{current_date}.xlsx')
wb.save(output_path)

print(f"Dashboard en data zijn opgeslagen in {output_path}")
